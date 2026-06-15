import os
import tempfile
from openpyxl import load_workbook
from openpyxl import Workbook


# =====================================================
# HELPERS
# =====================================================

def is_numeric(value):

    try:
        float(value)
        return True
    except:
        return False


def to_float(value):

    try:
        return float(value)
    except:
        return 0.0


def is_bold_row(row):

    bold_count = 0

    for cell in row:

        try:

            if cell.font and cell.font.bold:
                bold_count += 1

        except:
            pass

    return bold_count >= 2


def get_bold_rows(sheet):

    bold_rows = []

    for row_num in range(1, sheet.max_row + 1):

        row = list(sheet[row_num])

        if is_bold_row(row):
            bold_rows.append(row_num)

    return bold_rows


def get_numeric_columns(sheet):

    numeric_cols = set()

    for row in sheet.iter_rows():

        for cell in row:

            if is_numeric(cell.value):
                numeric_cols.add(cell.column)

    return sorted(list(numeric_cols))


# =====================================================
# VALIDATE SHEET
# =====================================================

def validate_sheet(sheet):

    mismatches = []

    bold_rows = get_bold_rows(sheet)

    if len(bold_rows) < 2:
        return mismatches

    numeric_columns = get_numeric_columns(sheet)

    for block_index in range(len(bold_rows) - 1):

        start_row = bold_rows[block_index]
        end_row = bold_rows[block_index + 1]

        block_name = str(
            sheet.cell(start_row, 1).value
        )

        for col in numeric_columns:

            expected = to_float(
                sheet.cell(start_row, col).value
            )

            actual = 0

            for r in range(start_row + 1, end_row):

                value = sheet.cell(r, col).value

                if is_numeric(value):
                    actual += to_float(value)

            difference = round(expected - actual, 2)

            if abs(difference) > 0.01:

                mismatches.append([
                    sheet.title,
                    block_name,
                    sheet.cell(1, col).value
                    if sheet.cell(1, col).value
                    else f"Column {col}",
                    expected,
                    round(actual, 2),
                    difference,
                    "Mismatch"
                ])

    return mismatches

# =====================================================
# MAIN PROCESS
# =====================================================

def process_salary_file(file_path):

    extension = os.path.splitext(
        file_path
    )[1].lower()

    if extension != ".xlsx":

        raise Exception(
            "Only .xlsx files are supported."
        )

    wb = load_workbook(file_path)

    all_mismatches = []

    for sheet in wb.worksheets:

        if sheet.title == "Validation_Report":
            continue

        sheet_mismatches = validate_sheet(
            sheet
        )

        all_mismatches.extend(
            sheet_mismatches
        )

    if "Validation_Report" in wb.sheetnames:

        del wb["Validation_Report"]

    report_sheet = wb.create_sheet(
        "Validation_Report"
    )

    report_sheet.append([
        "Sheet Name",
        "Block Name",
        "Column",
        "Expected Value",
        "Actual Sum",
        "Difference",
        "Status"
    ])

    for row in all_mismatches:

        report_sheet.append(
            row
        )

    output_dir = tempfile.mkdtemp()

    original_name = os.path.basename(
        file_path
    )

    output_file = os.path.join(
        output_dir,
        "Validated_" + original_name
    )

    wb.save(output_file)

    return output_file, len(all_mismatches)